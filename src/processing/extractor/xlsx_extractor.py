"""
Extractor for XLSX files using the openpyxl library.
"""
import openpyxl

from .base_extractor import BaseExtractor

class XlsxExtractor(BaseExtractor):
    """Extracts text content from XLSX files."""

    def extract(self, file_path: str) -> str:
        """
        Opens an XLSX file and extracts all text from all cells in all sheets.

        Args:
            file_path: The path to the XLSX file.

        Returns:
            The concatenated text content from all cells.
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            full_text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                full_text.append(f"--- Sheet: {sheet_name} ---\n")
                for row in sheet.iter_rows():
                    row_text = []
                    for cell in row:
                        if cell.value is not None:
                            row_text.append(str(cell.value))
                    if row_text:
                        full_text.append(" | ".join(row_text))
            return '\n'.join(full_text)
        except Exception as e:
            print(f"[ERROR] Failed to extract text from XLSX {file_path}: {e}")
            return "" # Return empty string on failure
